"""
dashboard.py — AMPIN_FTG_4 Dashboard
=====================================
Connects DIRECTLY to the three FTP servers (SECI-3, C&I, SECI-5) every time
it refreshes, finds the right report file(s) for the selected date RANGE,
downloads them INTO MEMORY (never saved to disk), and plots HYBRID, SOLAR,
and WIND active power together on one continuous chart per source spanning
the whole range.

No local copies of the FTP files are kept.

FILE STRUCTURE ON THE FTP SERVER (confirmed from a live FileZilla listing)
----------------------------------------------------------------------------
Each plant writes a NEW file every ~15 minutes, named like:

    FTP_REPORT 20-06-2026 12_00_00.xls
    FTP_REPORT 20-06-2026 11_45_00.xls
    ...

Each file contains EVERY row from the start of that day (00:15) up through
that file's own timestamp — i.e. the latest-timestamped file for a given
date already contains that entire day's data so far. For a date RANGE,
this script fetches the latest file for EACH day in the range and
concatenates them into one continuous series.

COLUMN LAYOUT PER SOURCE (by Excel column letter, NOT by header name —
confirmed directly against real files downloaded from each source's FTP
server on 2026-06-20.)
--------------------------------------------------------------------------
  C&I:     Timestamp=A  Hybrid=B  Wind=BF                Solar=BH
  SECI-3:  Timestamp=A  Hybrid=B  Wind=Hybrid-Solar (calc)  Solar=CQ+CU
  SECI-5:  Timestamp=B  Hybrid=C  Wind=CK                Solar=CM+CQ

TIMESTAMP DISPLAY OFFSET
---------------------------
Each row's timestamp is shifted back by 15 minutes for display purposes
(e.g. a row timestamped 00:15 in the file is shown as 00:00 on the chart).

SETUP
-----
1. Copy ".env.template" to ".env" in the same folder as this script and
   fill in/confirm the FTP credentials.
2. Install dependencies (one time):
       pip install flask pandas openpyxl xlrd python-dotenv
3. Run:
       python dashboard.py
4. Open http://127.0.0.1:5000 and leave the tab open — it auto-refreshes
   every POLL_INTERVAL_MINUTES (set in .env).
"""

import os
import re
import sys
import io
from pathlib import Path
from datetime import datetime, date, timedelta
from ftplib import FTP, error_perm

from flask import Flask, render_template_string, jsonify, request, send_file
import pandas as pd
import openpyxl.utils as xl_utils
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

try:
    from dotenv import load_dotenv
except ImportError:
    print("Missing dependency. Run:  pip install python-dotenv")
    sys.exit(1)

SCRIPT_DIR = Path(__file__).resolve().parent
ENV_PATH = SCRIPT_DIR / ".env"

# Load .env if it exists (local development). On hosts like Render, there is
# no .env file — credentials are injected directly as environment variables
# via the platform's dashboard, so this is skipped there and that's expected.
if ENV_PATH.exists():
    load_dotenv(ENV_PATH)

# Only hard-fail if NEITHER a .env file NOR real environment variables are
# present — i.e. genuinely unconfigured, whether running locally or deployed.
if not ENV_PATH.exists() and not os.getenv("SECI3_HOST"):
    print(
        "No FTP credentials found.\n"
        "  - Running locally? Copy .env.template to .env and fill it in.\n"
        "  - Running on Render (or similar)? Set the credentials as "
        "environment variables in the platform's dashboard instead."
    )
    sys.exit(1)


def env(key, default=None, required=True):
    val = os.getenv(key, default)
    if required and (val is None or val == ""):
        print(f"Missing required setting '{key}' in .env file.")
        sys.exit(1)
    return val


SOURCES = {
    "SECI-3": {
        "host": env("SECI3_HOST"),
        "port": int(env("SECI3_PORT", "21")),
        "user": env("SECI3_USER"),
        "password": env("SECI3_PASS"),
        "remote_dir": env("SECI3_REMOTE_DIR", "/"),
    },
    "CNI": {
        "host": env("CNI_HOST"),
        "port": int(env("CNI_PORT", "21")),
        "user": env("CNI_USER"),
        "password": env("CNI_PASS"),
        "remote_dir": env("CNI_REMOTE_DIR", "/"),
    },
    "SECI-5": {
        "host": env("SECI5_HOST"),
        "port": int(env("SECI5_PORT", "21")),
        "user": env("SECI5_USER"),
        "password": env("SECI5_PASS"),
        "remote_dir": env("SECI5_REMOTE_DIR", "/"),
    },
    "MSEDCL": {
        "host": env("MSEDCL_HOST"),
        "port": int(env("MSEDCL_PORT", "21")),
        "user": env("MSEDCL_USER"),
        "password": env("MSEDCL_PASS"),
        "remote_dir": env("MSEDCL_REMOTE_DIR", "/"),
    },
}

SOURCE_DISPLAY_NAMES = {
    "SECI-3": "SECI-3",
    "CNI": "C&I",
    "SECI-5": "SECI-5",
    "MSEDCL": "MSEDCL 200MW",
}

POLL_INTERVAL_MINUTES = float(env("POLL_INTERVAL_MINUTES", "15", required=False) or 15)

# Maximum days a single range query will fetch, to keep response times sane
# (each day = one FTP login + listing + download per source).
MAX_RANGE_DAYS = 31

SOURCE_COLUMNS = {
    "CNI": {
        "timestamp": "A",
        "hybrid": "B",
        "wind": "BF",
        "solar": "BH",
    },
    "SECI-3": {
        "timestamp": "A",
        "hybrid": "B",
        "solar": ["CQ", "CU"],
        "wind": "HYBRID_MINUS_SOLAR",
    },
    "SECI-5": {
        "timestamp": "B",
        "hybrid": "C",
        "wind": "CK",
        "solar": ["CM", "CQ"],
    },
}

TIMESTAMP_DISPLAY_OFFSET_MINUTES = 15

app = Flask(__name__)


# --------------------------------------------------------------------------
# Filename parsing
# --------------------------------------------------------------------------

FILENAME_PATTERN = re.compile(
    r"(\d{2}-\d{2}-\d{4})[ _](\d{1,2})_(\d{2})_(\d{2})"
)


def parse_filename(fname: str):
    m = FILENAME_PATTERN.search(fname)
    if not m:
        return None
    date_str, hh, mm, ss = m.groups()
    try:
        d = datetime.strptime(date_str, "%d-%m-%Y").date()
    except ValueError:
        return None
    return d, (int(hh), int(mm), int(ss))


def find_latest_file_for_date(filenames, target_date: date):
    candidates = []
    for f in filenames:
        parsed = parse_filename(f)
        if parsed and parsed[0] == target_date:
            candidates.append((f, parsed[1]))
    if not candidates:
        return None
    candidates.sort(key=lambda pair: pair[1])
    return candidates[-1][0]


def date_range_list(start: date, end: date):
    if end < start:
        start, end = end, start
    days = []
    d = start
    while d <= end and len(days) < MAX_RANGE_DAYS:
        days.append(d)
        d += timedelta(days=1)
    return days


# --------------------------------------------------------------------------
# FTP: list files + fetch one into memory
# --------------------------------------------------------------------------

def ftp_connect(source: dict) -> FTP:
    ftp = FTP()
    ftp.connect(source["host"], source["port"], timeout=30)
    ftp.login(source["user"], source["password"])
    ftp.set_pasv(True)
    ftp.cwd(source["remote_dir"])
    return ftp


def list_report_files(ftp: FTP):
    names = ftp.nlst()
    return [n for n in names if n.lower().endswith((".xls", ".xlsx"))]


def fetch_file_bytes(ftp: FTP, filename: str) -> bytes:
    buf = io.BytesIO()
    ftp.retrbinary(f"RETR {filename}", buf.write)
    return buf.getvalue()


# --------------------------------------------------------------------------
# Excel parsing (in-memory, column-letter based)
# --------------------------------------------------------------------------

def _engine_for(filename: str):
    return "xlrd" if filename.lower().endswith(".xls") else "openpyxl"


def _find_header_row(file_bytes: bytes, engine: str) -> int:
    preview = pd.read_excel(
        io.BytesIO(file_bytes), sheet_name=0, header=None, nrows=6, engine=engine
    )
    for idx, row in preview.iterrows():
        row_vals = [str(v).strip().upper() for v in row.tolist()]
        if any("TIMESTAMP" in v for v in row_vals):
            return idx
    return 1


def _col_index(letter: str) -> int:
    return xl_utils.column_index_from_string(letter) - 1


def _sum_columns(df: pd.DataFrame, letters):
    if isinstance(letters, str) and letters != "HYBRID_MINUS_SOLAR":
        letters = [letters]
    total = None
    for letter in letters:
        idx = _col_index(letter)
        if idx >= df.shape[1]:
            raise ValueError(
                f"Column '{letter}' (position {idx}) is out of range — "
                f"the file only has {df.shape[1]} columns."
            )
        series = pd.to_numeric(df.iloc[:, idx], errors="coerce")
        total = series if total is None else (total + series)
    return total


def parse_report_df(file_bytes: bytes, filename: str, source_name: str) -> pd.DataFrame:
    """Returns a cleaned DataFrame with columns: timestamp, display_timestamp,
    hybrid_mw, solar_mw, wind_mw — for ONE file."""
    engine = _engine_for(filename)
    header_row = _find_header_row(file_bytes, engine)
    df = pd.read_excel(io.BytesIO(file_bytes), sheet_name=0, header=None, engine=engine)
    df = df.iloc[header_row + 1:].reset_index(drop=True)

    cols = SOURCE_COLUMNS[source_name]

    ts_idx = _col_index(cols["timestamp"])
    if ts_idx >= df.shape[1]:
        raise ValueError(f"Timestamp column '{cols['timestamp']}' out of range.")
    timestamps = pd.to_datetime(df.iloc[:, ts_idx], errors="coerce")

    hybrid_idx = _col_index(cols["hybrid"])
    if hybrid_idx >= df.shape[1]:
        raise ValueError(f"Hybrid column '{cols['hybrid']}' out of range.")
    hybrid = pd.to_numeric(df.iloc[:, hybrid_idx], errors="coerce")

    solar = _sum_columns(df, cols["solar"])

    if cols["wind"] == "HYBRID_MINUS_SOLAR":
        wind = hybrid - solar
    else:
        wind = _sum_columns(df, cols["wind"])

    clean = pd.DataFrame({
        "timestamp": timestamps,
        "hybrid_mw": hybrid,
        "solar_mw": solar,
        "wind_mw": wind,
    })
    clean = clean.dropna(subset=["timestamp"])
    clean = clean.sort_values("timestamp")
    clean["display_timestamp"] = clean["timestamp"] - timedelta(
        minutes=TIMESTAMP_DISPLAY_OFFSET_MINUTES
    )
    return clean


# --------------------------------------------------------------------------
# Core: fetch + parse for one source across a date RANGE
# --------------------------------------------------------------------------

def load_source_range(source_name: str, from_date: date, to_date: date):
    """
    Fetches the latest file for each day in [from_date, to_date], parses
    each, and concatenates into one continuous DataFrame (sorted by time).
    """
    source = SOURCES[source_name]
    days = date_range_list(from_date, to_date)
    if not days:
        return {"ok": False, "error": "Invalid date range."}

    try:
        ftp = ftp_connect(source)
    except Exception as e:
        return {"ok": False, "error": f"Could not connect/login to {source_name}: {e}"}

    frames = []
    files_used = []
    warnings = []

    try:
        try:
            filenames = list_report_files(ftp)
        except error_perm as e:
            return {"ok": False, "error": f"Cannot list remote directory: {e}"}

        if not filenames:
            return {"ok": False, "error": "No .xls/.xlsx files found on server."}

        for d in days:
            match = find_latest_file_for_date(filenames, d)
            if not match:
                warnings.append(f"No file found for {d.strftime('%d-%m-%Y')}")
                continue
            try:
                file_bytes = fetch_file_bytes(ftp, match)
                day_df = parse_report_df(file_bytes, match, source_name)
                frames.append(day_df)
                files_used.append(match)
            except Exception as e:
                warnings.append(f"{match}: {e}")

    finally:
        try:
            ftp.quit()
        except Exception:
            ftp.close()

    if not frames:
        msg = "No data could be retrieved for the selected range."
        if warnings:
            msg += " " + "; ".join(warnings)
        return {"ok": False, "error": msg}

    combined = pd.concat(frames, ignore_index=True)
    combined = combined.drop_duplicates(subset=["timestamp"]).sort_values("timestamp")

    return {
        "ok": True,
        "error": None,
        "files_used": files_used,
        "warnings": warnings,
        "from_date": from_date.isoformat(),
        "to_date": to_date.isoformat(),
        "fetched_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "timestamps": combined["display_timestamp"].dt.strftime("%Y-%m-%dT%H:%M:%S").tolist(),
        "hybrid": combined["hybrid_mw"].round(3).tolist(),
        "solar": combined["solar_mw"].round(3).tolist(),
        "wind": combined["wind_mw"].round(3).tolist(),
        "_df": combined,  # kept for export use; stripped before jsonify
    }


# Sources that use XLS files (FTP_REPORT DD-MM-YYYY HH_MM_SS.xls pattern)
XLS_SOURCES = {"SECI-3", "CNI", "SECI-5"}

# Sources that use CSV files (DateDD-MM-YYYY.csv pattern, one file per day)
CSV_SOURCES = {"MSEDCL"}


# --------------------------------------------------------------------------
# MSEDCL: CSV fetch + parse (completely separate from the XLS pipeline)
# --------------------------------------------------------------------------

def msedcl_filename_for_date(target_date: date) -> str:
    """Returns the expected CSV filename for a given date."""
    return f"Date{target_date.strftime('%d-%m-%Y')}.csv"


def parse_msedcl_csv(file_bytes: bytes) -> pd.DataFrame:
    """
    Parse MSEDCL's daily CSV file into a DataFrame with columns:
        timestamp, display_timestamp, plant_active_power_mw
    File is latin-1 encoded (has degree symbol ° in a header), uses
    Windows line endings (\r\n), and has no extra header/title row —
    the very first line is the column header.
    """
    df = pd.read_csv(
        io.BytesIO(file_bytes),
        encoding="latin-1",
        sep=",",
        skipinitialspace=True,
    )
    # Strip \r and whitespace from column names (Windows line endings)
    df.columns = [c.strip().replace('\r', '') for c in df.columns]

    ts_col = "DateAndTime"
    power_col = "Plant Active Power(MW)"

    if ts_col not in df.columns:
        raise ValueError(f"Timestamp column '{ts_col}' not found in CSV.")
    if power_col not in df.columns:
        raise ValueError(f"Column '{power_col}' not found in CSV.")

    df[ts_col] = df[ts_col].astype(str).str.strip().str.replace('\r', '')
    timestamps = pd.to_datetime(df[ts_col], format="%d-%m-%Y %H:%M", errors="coerce")
    power = pd.to_numeric(df[power_col], errors="coerce")

    clean = pd.DataFrame({
        "timestamp": timestamps,
        "plant_active_power_mw": power,
    })
    clean = clean.dropna(subset=["timestamp"]).sort_values("timestamp")
    clean["display_timestamp"] = clean["timestamp"] - timedelta(
        minutes=TIMESTAMP_DISPLAY_OFFSET_MINUTES
    )
    return clean


def load_msedcl_range(from_date: date, to_date: date):
    """
    For MSEDCL: fetches DateDD-MM-YYYY.csv for each day in [from_date, to_date],
    parses each, and concatenates into one continuous series.
    """
    source = SOURCES["MSEDCL"]
    days = date_range_list(from_date, to_date)
    if not days:
        return {"ok": False, "error": "Invalid date range."}

    try:
        ftp = ftp_connect(source)
    except Exception as e:
        return {"ok": False, "error": f"Could not connect/login to MSEDCL: {e}"}

    frames = []
    files_used = []
    warnings = []

    try:
        try:
            all_files = ftp.nlst()
        except error_perm as e:
            return {"ok": False, "error": f"Cannot list remote directory: {e}"}

        for d in days:
            fname = msedcl_filename_for_date(d)
            if fname not in all_files:
                warnings.append(f"No file found for {d.strftime('%d-%m-%Y')} (expected '{fname}')")
                continue
            try:
                file_bytes = fetch_file_bytes(ftp, fname)
                day_df = parse_msedcl_csv(file_bytes)
                frames.append(day_df)
                files_used.append(fname)
            except Exception as e:
                warnings.append(f"{fname}: {e}")

    finally:
        try:
            ftp.quit()
        except Exception:
            ftp.close()

    if not frames:
        msg = "No data could be retrieved for the selected range."
        if warnings:
            msg += " " + "; ".join(warnings)
        return {"ok": False, "error": msg}

    combined = pd.concat(frames, ignore_index=True)
    combined = combined.drop_duplicates(subset=["timestamp"]).sort_values("timestamp")

    return {
        "ok": True,
        "error": None,
        "files_used": files_used,
        "warnings": warnings,
        "from_date": from_date.isoformat(),
        "to_date": to_date.isoformat(),
        "fetched_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "timestamps": combined["display_timestamp"].dt.strftime("%Y-%m-%dT%H:%M:%S").tolist(),
        "plant_power": combined["plant_active_power_mw"].round(3).tolist(),
        "_df": combined,
        "_type": "msedcl",
    }


# --------------------------------------------------------------------------
# Diagnostics
# --------------------------------------------------------------------------

@app.route("/api/debug-columns")
def api_debug_columns():
    src = request.args.get("source")
    if src not in SOURCES:
        return jsonify({"error": f"source must be one of {list(SOURCES.keys())}"}), 400

    date_param = request.args.get("date", date.today().isoformat())
    try:
        target_date = datetime.strptime(date_param, "%Y-%m-%d").date()
    except ValueError:
        target_date = date.today()

    source = SOURCES[src]
    try:
        ftp = ftp_connect(source)
    except Exception as e:
        return jsonify({"error": f"Could not connect/login: {e}"}), 200

    try:
        filenames = list_report_files(ftp)
        match = find_latest_file_for_date(filenames, target_date)
        if not match:
            return jsonify({
                "error": f"No file matched for {target_date}",
                "all_files_on_server": filenames,
            }), 200
        file_bytes = fetch_file_bytes(ftp, match)
    finally:
        try:
            ftp.quit()
        except Exception:
            ftp.close()

    engine = _engine_for(match)
    try:
        header_row = _find_header_row(file_bytes, engine)
        raw = pd.read_excel(io.BytesIO(file_bytes), sheet_name=0, header=None, engine=engine)
        header_text = raw.iloc[header_row].tolist()
        first_data_row = raw.iloc[header_row + 1].tolist()
        lettered_headers = {
            xl_utils.get_column_letter(i + 1): str(v)
            for i, v in enumerate(header_text) if pd.notna(v)
        }
    except Exception as e:
        return jsonify({"error": f"Failed to parse: {e}", "matched_file": match}), 200

    return jsonify({
        "matched_file": match,
        "header_row_index": header_row,
        "headers_by_column_letter": lettered_headers,
        "first_data_row_by_column_letter": {
            xl_utils.get_column_letter(i + 1): v
            for i, v in enumerate(first_data_row)
        },
    })


# --------------------------------------------------------------------------
# Routes
# --------------------------------------------------------------------------

def _parse_range_params(src: str):
    today_str = date.today().isoformat()
    from_param = request.args.get(f"from_{src}", today_str)
    to_param = request.args.get(f"to_{src}", today_str)
    try:
        from_date = datetime.strptime(from_param, "%Y-%m-%d").date()
    except ValueError:
        from_date = date.today()
    try:
        to_date = datetime.strptime(to_param, "%Y-%m-%d").date()
    except ValueError:
        to_date = date.today()
    return from_date, to_date


@app.route("/api/data")
def api_data():
    result = {}
    for src in SOURCES:
        from_date, to_date = _parse_range_params(src)
        if src in CSV_SOURCES:
            data = load_msedcl_range(from_date, to_date)
        else:
            data = load_source_range(src, from_date, to_date)
        data.pop("_df", None)
        result[src] = data

    result["_meta"] = {
        "refresh_ms": int(POLL_INTERVAL_MINUTES * 60 * 1000),
        "server_time": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "today": date.today().isoformat(),
    }
    return jsonify(result)


@app.route("/api/export")
def api_export():
    """
    Builds an in-memory .xlsx with one sheet per source (SECI-3, C&I,
    SECI-5), using each source's currently-selected date range, and
    returns it as a file download.
    """
    wb = Workbook()
    wb.remove(wb.active)  # remove default empty sheet

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", start_color="2F3A56")
    header_align = Alignment(horizontal="center")

    any_data = False

    for src in SOURCES:
        from_date, to_date = _parse_range_params(src)
        if src in CSV_SOURCES:
            data = load_msedcl_range(from_date, to_date)
        else:
            data = load_source_range(src, from_date, to_date)

        sheet_title = SOURCE_DISPLAY_NAMES[src][:31]
        ws = wb.create_sheet(title=sheet_title)

        if not data.get("ok"):
            ws.append(["Error", data.get("error", "Unknown error")])
            continue

        any_data = True

        if data.get("_type") == "msedcl":
            df = data["_df"][["display_timestamp", "plant_active_power_mw"]].copy()
            df.columns = ["Timestamp", "Plant Active Power (MW)"]
            col_count = 2
        else:
            df = data["_df"][["display_timestamp", "hybrid_mw", "solar_mw", "wind_mw"]].copy()
            df.columns = [
                "Timestamp",
                "Hybrid Active Power (MW)",
                "Solar Active Power (MW)",
                "Wind Active Power (MW)",
            ]
            col_count = 4

        ws.append([f"{SOURCE_DISPLAY_NAMES[src]} — {from_date} to {to_date}"])
        ws.append([])
        header_row_idx = ws.max_row + 1
        for row in dataframe_to_rows(df, index=False, header=True):
            ws.append(row)

        for col_idx in range(1, col_count + 1):
            cell = ws.cell(row=header_row_idx, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align

        ws.column_dimensions["A"].width = 18
        for letter in ["B", "C", "D"][:col_count - 1]:
            ws.column_dimensions[letter].width = 26

    if not any_data:
        ws = wb.create_sheet(title="No Data")
        ws.append(["No data was available for any source/date range selected."])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"AMPIN_FTG_4_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(
        buf,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


PAGE_TEMPLATE = """
<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>AMPIN_FTG_4 Dashboard</title>
<script src="https://cdn.plot.ly/plotly-2.32.0.min.js"></script>
<link rel="preconnect" href="https://fonts.googleapis.com">
<link href="https://fonts.googleapis.com/css2?family=Space+Grotesk:wght@500;600;700&family=Inter:wght@400;500;600;700&family=JetBrains+Mono:wght@400;500&display=swap" rel="stylesheet">
<style>
  :root {
    /* Base */
    --bg: #f6f7f5;
    --surface: #ffffff;
    --surface-2: #fbfbfa;
    --border: #e4e7e2;
    --border-soft: #edefea;
    --text: #16201c;
    --muted: #6b756f;
    --muted-2: #99a39c;

    /* Signal colors — drawn from the plant's own world: sun, sky/air, and
       the merged grid output. Not generic chart-library blue/red/teal. */
    --solar: #e8910c;
    --solar-soft: #fdf1de;
    --wind: #0f9e95;
    --wind-soft: #e1f5f2;
    --hybrid: #3d3a8c;
    --hybrid-soft: #ebebf7;

    --good: #1ea672;
    --bad: #d23c4f;

    --shadow-sm: 0 1px 2px rgba(22,32,28,0.04), 0 1px 1px rgba(22,32,28,0.03);
    --shadow-md: 0 6px 20px rgba(22,32,28,0.06), 0 2px 6px rgba(22,32,28,0.04);
    --shadow-lift: 0 14px 32px rgba(22,32,28,0.10), 0 4px 10px rgba(22,32,28,0.05);
  }

  * { box-sizing: border-box; }

  body {
    margin: 0;
    background:
      radial-gradient(1200px 480px at 8% -10%, #fff7e8 0%, rgba(255,247,232,0) 60%),
      radial-gradient(1000px 420px at 100% 0%, #e7f7f4 0%, rgba(231,247,244,0) 55%),
      var(--bg);
    color: var(--text);
    font-family: "Inter", -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, Arial, sans-serif;
    padding: 28px 36px 56px;
    -webkit-font-smoothing: antialiased;
  }

  /* ---------- Top bar ---------- */
  .top-bar {
    display: flex;
    justify-content: space-between;
    align-items: flex-end;
    flex-wrap: wrap;
    gap: 18px;
    margin-bottom: 26px;
    padding-bottom: 22px;
    border-bottom: 1px solid var(--border-soft);
  }
  .brand-row { display: flex; align-items: center; gap: 12px; }
  .brand-mark {
    width: 38px; height: 38px;
    border-radius: 10px;
    background: conic-gradient(from 220deg, var(--solar), var(--hybrid) 50%, var(--wind));
    box-shadow: var(--shadow-sm);
    position: relative;
    flex-shrink: 0;
  }
  .brand-mark::after {
    content: "";
    position: absolute; inset: 4px;
    border-radius: 7px;
    background: var(--surface);
  }
  h1 {
    font-family: "Space Grotesk", "Inter", sans-serif;
    font-size: 25px;
    font-weight: 700;
    margin: 0 0 3px;
    letter-spacing: -0.02em;
    color: var(--text);
  }
  .subtitle {
    color: var(--muted);
    font-size: 13px;
    margin-bottom: 2px;
  }
  .subtitle .dim { color: var(--muted-2); }
  .subtitle b.s { color: var(--solar); font-weight: 600; }
  .subtitle b.w { color: var(--wind); font-weight: 600; }
  .subtitle b.h { color: var(--hybrid); font-weight: 600; }

  .download-btn {
    display: inline-flex;
    align-items: center;
    gap: 9px;
    background: var(--text);
    border: 1px solid var(--text);
    color: #fff;
    border-radius: 10px;
    padding: 11px 18px;
    font-family: "Inter", sans-serif;
    font-size: 13.5px;
    font-weight: 600;
    cursor: pointer;
    box-shadow: var(--shadow-sm);
    white-space: nowrap;
    transition: transform 0.15s ease, box-shadow 0.15s ease, background 0.15s ease;
  }
  .download-btn:hover {
    background: #2b3933;
    box-shadow: var(--shadow-md);
    transform: translateY(-1px);
  }
  .download-btn:active { transform: translateY(0); box-shadow: var(--shadow-sm); }
  .download-btn svg { width: 15px; height: 15px; }

  /* ---------- Status strip ---------- */
  .status-bar {
    display: flex;
    justify-content: space-between;
    align-items: center;
    margin-bottom: 22px;
    font-size: 12.5px;
    color: var(--muted);
    flex-wrap: wrap;
    gap: 8px;
  }
  .live-chip {
    display: inline-flex;
    align-items: center;
    gap: 7px;
    background: var(--surface);
    border: 1px solid var(--border-soft);
    border-radius: 999px;
    padding: 5px 12px 5px 9px;
    box-shadow: var(--shadow-sm);
  }
  .dot {
    display: inline-block;
    width: 7px; height: 7px;
    border-radius: 50%;
    background: var(--good);
    box-shadow: 0 0 0 3px rgba(30,166,114,0.16);
    animation: pulse 2.2s infinite;
  }
  @keyframes pulse {
    0%, 100% { box-shadow: 0 0 0 3px rgba(30,166,114,0.16); }
    50% { box-shadow: 0 0 0 6px rgba(30,166,114,0.05); }
  }

  /* ---------- Grid / cards ---------- */
  .grid { display: grid; grid-template-columns: 1fr; gap: 20px; }

  .panel {
    background: var(--surface);
    border: 1px solid var(--border-soft);
    border-radius: 16px;
    padding: 22px 24px 18px;
    box-shadow: var(--shadow-sm);
    transition: box-shadow 0.25s ease, transform 0.25s ease;
    position: relative;
    overflow: hidden;
  }
  .panel:hover {
    box-shadow: var(--shadow-lift);
    transform: translateY(-2px);
  }
  .panel::before {
    content: "";
    position: absolute;
    top: 0; left: 0; right: 0;
    height: 3px;
    background: var(--accent-grad, linear-gradient(90deg, var(--hybrid), var(--wind)));
  }
  .panel[data-plant="SECI-3"]::before { background: linear-gradient(90deg, #3d3a8c, #6e6bd1); }
  .panel[data-plant="CNI"]::before { background: linear-gradient(90deg, #0f9e95, #4fd6cb); }
  .panel[data-plant="SECI-5"]::before { background: linear-gradient(90deg, #e8910c, #f7b94d); }
  .panel[data-plant="MSEDCL"]::before { background: linear-gradient(90deg, #7c3aed, #a78bfa); }

  .panel-header {
    display: flex;
    justify-content: space-between;
    align-items: center;
    margin-bottom: 14px;
    flex-wrap: wrap;
    gap: 12px;
  }
  .panel-title-group { display: flex; align-items: center; gap: 12px; flex-wrap: wrap; }
  .panel-title {
    font-family: "Space Grotesk", sans-serif;
    font-size: 16px;
    font-weight: 700;
    letter-spacing: -0.01em;
  }
  .panel[data-plant="SECI-3"] .panel-title { color: #3d3a8c; }
  .panel[data-plant="CNI"] .panel-title { color: #0f9e95; }
  .panel[data-plant="SECI-5"] .panel-title { color: #c97509; }
  .panel[data-plant="MSEDCL"] .panel-title { color: #7c3aed; }

  .panel-meta {
    font-family: "JetBrains Mono", monospace;
    font-size: 11px;
    color: var(--muted-2);
    text-align: right;
  }

  .range-picker-wrap {
    display: flex;
    align-items: center;
    gap: 7px;
    font-size: 12px;
    color: var(--muted);
    flex-wrap: wrap;
  }
  .range-picker-wrap label { font-weight: 500; }
  .range-picker-wrap input[type="date"] {
    background: var(--surface-2);
    border: 1px solid var(--border);
    color: var(--text);
    border-radius: 8px;
    padding: 6px 9px;
    font-size: 12.5px;
    font-family: "Inter", sans-serif;
    transition: border-color 0.15s ease;
  }
  .range-picker-wrap input[type="date"]:hover,
  .range-picker-wrap input[type="date"]:focus {
    border-color: var(--muted-2);
    outline: none;
  }
  .today-btn {
    background: var(--surface-2);
    border: 1px solid var(--border);
    color: var(--text);
    border-radius: 8px;
    padding: 6px 13px;
    font-size: 12px;
    font-weight: 600;
    cursor: pointer;
    transition: background 0.15s ease, border-color 0.15s ease;
  }
  .today-btn:hover { background: var(--surface); border-color: var(--muted-2); }

  /* ---------- Legend chips (live values) ---------- */
  .stat-row {
    display: flex;
    gap: 10px;
    margin: 2px 0 14px;
    flex-wrap: wrap;
  }
  .stat-chip {
    display: flex;
    align-items: baseline;
    gap: 7px;
    background: var(--surface-2);
    border: 1px solid var(--border-soft);
    border-radius: 10px;
    padding: 8px 12px;
    font-family: "JetBrains Mono", monospace;
  }
  .stat-chip .swatch { width: 8px; height: 8px; border-radius: 2px; flex-shrink: 0; }
  .stat-chip .label { font-family: "Inter", sans-serif; font-size: 11px; color: var(--muted); font-weight: 500; }
  .stat-chip .value { font-size: 13.5px; font-weight: 600; color: var(--text); }

  .error-box {
    color: var(--bad);
    font-size: 13px;
    padding: 16px;
    background: #fdedef;
    border: 1px solid #f6cdd3;
    border-radius: 10px;
    font-family: "Inter", sans-serif;
  }

  .chart { width: 100%; height: 400px; }

  @media (max-width: 640px) {
    body { padding: 18px 16px 40px; }
    .panel { padding: 18px 16px 14px; }
  }
</style>
</head>
<body>

  <div class="top-bar">
    <div>
      <div class="brand-row">
        <div class="brand-mark"></div>
        <div>
          <h1>AMPIN_FTG_4 Dashboard</h1>
          <div class="subtitle">
            <b class="h">Hybrid</b> &middot; <b class="s">Solar</b> &middot; <b class="w">Wind</b> Active Power
            <span class="dim">&mdash; fetched live from FTP, no local files saved</span>
          </div>
        </div>
      </div>
    </div>
    <button class="download-btn" id="download-btn">
      <svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M21 15v4a2 2 0 0 1-2 2H5a2 2 0 0 1-2-2v-4"/><polyline points="7 10 12 15 17 10"/><line x1="12" y1="15" x2="12" y2="3"/></svg>
      Download Excel
    </button>
  </div>

  <div class="status-bar">
    <div class="live-chip"><span class="dot"></span><span id="status-text">Loading…</span></div>
    <div class="subtitle dim" style="margin:0;">Times shown are 15 minutes earlier than the raw file timestamp</div>
  </div>

  <div class="grid" id="charts-grid">
    <div class="panel" data-plant="SECI-3">
      <div class="panel-header">
        <div class="panel-title-group">
          <div class="panel-title">SECI-3</div>
          <div class="range-picker-wrap">
            <label>From</label>
            <input type="date" id="from-SECI-3" data-src="SECI-3">
            <label>To</label>
            <input type="date" id="to-SECI-3" data-src="SECI-3">
            <button class="today-btn" data-src="SECI-3">Today</button>
          </div>
        </div>
        <div class="panel-meta" id="meta-SECI-3"></div>
      </div>
      <div class="stat-row" id="stats-SECI-3"></div>
      <div id="chart-SECI-3" class="chart"></div>
    </div>

    <div class="panel" data-plant="CNI">
      <div class="panel-header">
        <div class="panel-title-group">
          <div class="panel-title">C&amp;I</div>
          <div class="range-picker-wrap">
            <label>From</label>
            <input type="date" id="from-CNI" data-src="CNI">
            <label>To</label>
            <input type="date" id="to-CNI" data-src="CNI">
            <button class="today-btn" data-src="CNI">Today</button>
          </div>
        </div>
        <div class="panel-meta" id="meta-CNI"></div>
      </div>
      <div class="stat-row" id="stats-CNI"></div>
      <div id="chart-CNI" class="chart"></div>
    </div>

    <div class="panel" data-plant="SECI-5">
      <div class="panel-header">
        <div class="panel-title-group">
          <div class="panel-title">SECI-5</div>
          <div class="range-picker-wrap">
            <label>From</label>
            <input type="date" id="from-SECI-5" data-src="SECI-5">
            <label>To</label>
            <input type="date" id="to-SECI-5" data-src="SECI-5">
            <button class="today-btn" data-src="SECI-5">Today</button>
          </div>
        </div>
        <div class="panel-meta" id="meta-SECI-5"></div>
      </div>
      <div class="stat-row" id="stats-SECI-5"></div>
      <div id="chart-SECI-5" class="chart"></div>
    </div>

    <div class="panel" data-plant="MSEDCL">
      <div class="panel-header">
        <div class="panel-title-group">
          <div class="panel-title">MSEDCL 200MW</div>
          <div class="range-picker-wrap">
            <label>From</label>
            <input type="date" id="from-MSEDCL" data-src="MSEDCL">
            <label>To</label>
            <input type="date" id="to-MSEDCL" data-src="MSEDCL">
            <button class="today-btn" data-src="MSEDCL">Today</button>
          </div>
        </div>
        <div class="panel-meta" id="meta-MSEDCL"></div>
      </div>
      <div class="stat-row" id="stats-MSEDCL"></div>
      <div id="chart-MSEDCL" class="chart"></div>
    </div>
  </div>

<script>
const SOURCES = ["SECI-3", "CNI", "SECI-5", "MSEDCL"];
const COLORS = {
  hybrid: "#3d3a8c",
  hybridSoft: "rgba(61,58,140,0.10)",
  solar: "#e8910c",
  solarSoft: "rgba(232,145,12,0.10)",
  wind: "#0f9e95",
  windSoft: "rgba(15,158,149,0.10)",
  plantPower: "#7c3aed",
  plantPowerSoft: "rgba(124,58,237,0.10)"
};
let selectedRanges = {};
let refreshTimer = null;

function todayStr() {
  const d = new Date();
  return d.getFullYear() + "-" + String(d.getMonth()+1).padStart(2,"0") + "-" + String(d.getDate()).padStart(2,"0");
}

function lastVal(arr) {
  if (!arr || !arr.length) return null;
  for (let i = arr.length - 1; i >= 0; i--) {
    if (arr[i] !== null && arr[i] !== undefined && !Number.isNaN(arr[i])) return arr[i];
  }
  return null;
}

function renderStats(elId, data) {
  const el = document.getElementById(elId);
  if (!data.ok) { el.innerHTML = ""; return; }
  if (data.plant_power !== undefined) {
    // MSEDCL: single series
    const p = lastVal(data.plant_power);
    const fmt = (v) => v === null ? "—" : v.toFixed(2) + " MW";
    el.innerHTML = `
      <div class="stat-chip"><span class="swatch" style="background:${COLORS.plantPower}"></span><span class="label">Plant Active Power (latest)</span><span class="value">${fmt(p)}</span></div>
    `;
  } else {
    // XLS sources: Hybrid / Solar / Wind
    const h = lastVal(data.hybrid), s = lastVal(data.solar), w = lastVal(data.wind);
    const fmt = (v) => v === null ? "—" : v.toFixed(2) + " MW";
    el.innerHTML = `
      <div class="stat-chip"><span class="swatch" style="background:${COLORS.hybrid}"></span><span class="label">Hybrid (latest)</span><span class="value">${fmt(h)}</span></div>
      <div class="stat-chip"><span class="swatch" style="background:${COLORS.solar}"></span><span class="label">Solar (latest)</span><span class="value">${fmt(s)}</span></div>
      <div class="stat-chip"><span class="swatch" style="background:${COLORS.wind}"></span><span class="label">Wind (latest)</span><span class="value">${fmt(w)}</span></div>
    `;
  }
}

function renderChart(elId, data) {
  const el = document.getElementById(elId);
  if (!data.ok) {
    el.innerHTML = `<div class="error-box">${data.error}</div>`;
    return;
  }

  let traces;

  if (data.plant_power !== undefined) {
    // MSEDCL: single series — Plant Active Power(MW)
    traces = [{
      x: data.timestamps, y: data.plant_power,
      name: "Plant Active Power [MW]",
      mode: "lines+markers",
      line: { color: COLORS.plantPower, width: 2.75, shape: "spline", smoothing: 0.6 },
      marker: { size: 5, color: "#ffffff", line: { color: COLORS.plantPower, width: 2 } },
      fill: "tozeroy",
      fillcolor: COLORS.plantPowerSoft,
      hovertemplate: "<b>%{y:.2f} MW</b><extra>Plant Active Power</extra>"
    }];
  } else {
    // SECI-3 / CNI / SECI-5: three series
    traces = [
      {
        x: data.timestamps, y: data.hybrid,
        name: "Hybrid Active Power [MW]",
        mode: "lines+markers",
        line: { color: COLORS.hybrid, width: 2.75, shape: "spline", smoothing: 0.6 },
        marker: { size: 5, color: "#ffffff", line: { color: COLORS.hybrid, width: 2 } },
        fill: "tozeroy",
        fillcolor: COLORS.hybridSoft,
        hovertemplate: "<b>%{y:.2f} MW</b><extra>Hybrid</extra>"
      },
      {
        x: data.timestamps, y: data.solar,
        name: "Solar Active Power [MW]",
        mode: "lines+markers",
        line: { color: COLORS.solar, width: 2.25, shape: "spline", smoothing: 0.6 },
        marker: { size: 4.5, color: "#ffffff", line: { color: COLORS.solar, width: 2 } },
        hovertemplate: "<b>%{y:.2f} MW</b><extra>Solar</extra>"
      },
      {
        x: data.timestamps, y: data.wind,
        name: "Wind Active Power [MW]",
        mode: "lines+markers",
        line: { color: COLORS.wind, width: 2.25, shape: "spline", smoothing: 0.6 },
        marker: { size: 4.5, color: "#ffffff", line: { color: COLORS.wind, width: 2 } },
        hovertemplate: "<b>%{y:.2f} MW</b><extra>Wind</extra>"
      }
    ];
  }

  const layout = {
    margin: { t: 8, r: 16, l: 46, b: 36 },
    paper_bgcolor: "transparent",
    plot_bgcolor: "transparent",
    font: { color: "#16201c", size: 12, family: "Inter, sans-serif" },
    xaxis: {
      type: "date",
      gridcolor: "#eef0eb", title: "", zerolinecolor: "#eef0eb",
      showspikes: true, spikemode: "across", spikecolor: "#c7ccc4", spikethickness: 1, spikedash: "dot",
      tickfont: { color: "#6b756f" }
    },
    yaxis: {
      gridcolor: "#eef0eb", title: "MW", zerolinecolor: "#dfe3dc",
      tickfont: { color: "#6b756f" },
      titlefont: { color: "#6b756f", size: 11 }
    },
    legend: {
      orientation: "h", y: -0.16, x: 0.5, xanchor: "center",
      font: { color: "#4a534d", size: 11.5 }
    },
    hovermode: "x unified",
    hoverlabel: { bgcolor: "#16201c", bordercolor: "#16201c", font: { color: "#ffffff", size: 12 } },
    transition: { duration: 350, easing: "cubic-in-out" }
  };

  Plotly.newPlot(el, traces, layout, {
    responsive: true,
    displayModeBar: false
  });
}

async function refreshData() {
  try {
    const params = new URLSearchParams();
    for (const src of SOURCES) {
      const range = selectedRanges[src] || { from: todayStr(), to: todayStr() };
      params.set("from_" + src, range.from);
      params.set("to_" + src, range.to);
    }
    const res = await fetch("/api/data?" + params.toString());
    const data = await res.json();

    document.getElementById("status-text").textContent =
      "Live — last checked " + data._meta.server_time;

    for (const src of SOURCES) {
      renderChart("chart-" + src, data[src]);
      renderStats("stats-" + src, data[src]);
      const metaEl = document.getElementById("meta-" + src);
      if (data[src].ok) {
        const fileCount = (data[src].files_used || []).length;
        metaEl.textContent = `${fileCount} day(s) · ${data[src].fetched_at}`;
      } else {
        metaEl.textContent = "";
      }
    }

    if (refreshTimer) clearTimeout(refreshTimer);
    refreshTimer = setTimeout(refreshData, data._meta.refresh_ms);
  } catch (e) {
    document.getElementById("status-text").textContent = "Error fetching data: " + e;
    if (refreshTimer) clearTimeout(refreshTimer);
    refreshTimer = setTimeout(refreshData, 60000);
  }
}

for (const src of SOURCES) {
  selectedRanges[src] = { from: todayStr(), to: todayStr() };
  const fromInput = document.getElementById("from-" + src);
  const toInput = document.getElementById("to-" + src);
  fromInput.value = todayStr();
  toInput.value = todayStr();
  fromInput.max = todayStr();
  toInput.max = todayStr();

  fromInput.addEventListener("change", (e) => {
    selectedRanges[src].from = e.target.value;
    refreshData();
  });
  toInput.addEventListener("change", (e) => {
    selectedRanges[src].to = e.target.value;
    refreshData();
  });
}

document.querySelectorAll(".today-btn").forEach(btn => {
  btn.addEventListener("click", () => {
    const src = btn.dataset.src;
    selectedRanges[src] = { from: todayStr(), to: todayStr() };
    document.getElementById("from-" + src).value = todayStr();
    document.getElementById("to-" + src).value = todayStr();
    refreshData();
  });
});

document.getElementById("download-btn").addEventListener("click", () => {
  const params = new URLSearchParams();
  for (const src of SOURCES) {
    const range = selectedRanges[src] || { from: todayStr(), to: todayStr() };
    params.set("from_" + src, range.from);
    params.set("to_" + src, range.to);
  }
  window.location.href = "/api/export?" + params.toString();
});

refreshData();
</script>

</body>
</html>
"""


@app.route("/")
def index():
    return render_template_string(PAGE_TEMPLATE)


if __name__ == "__main__":
    # On Render (and most cloud hosts), the PORT env var is set automatically
    # and the app must bind to 0.0.0.0 to be reachable externally. Locally,
    # with no PORT set, default to 127.0.0.1:5000 as before.
    port_env = os.getenv("PORT")
    if port_env:
        host = "0.0.0.0"
        port = int(port_env)
    else:
        host = os.getenv("HOST", "127.0.0.1")
        port = int(os.getenv("PORT", "5000"))
    print(f"Starting dashboard, binding to {host}:{port}")
    print(f"Auto-refresh every {POLL_INTERVAL_MINUTES} minute(s). Reading live from FTP — no local files saved.")
    app.run(host=host, port=port, debug=False)